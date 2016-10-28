import openpyxl
import os
import sys
import win32com.client
from time import sleep
from tkinter.messagebox import askyesnocancel
import platform

debug_it = 0

email_part_1 = """Greetings,

You are listed as delivery contact for site """

email_part_2 = """.  Can you please check the following and notify us of any corrections:

Delivery Info:

"""

email_part_3 = """

You may receive multiple emails if there are variations in the details for this site.  Please let us know the most accurate information so we can update our records.

Thank you,
"""


def send_mail_via_com(text, subject, recipient):
    olMailItem = 0x0
    obj = win32com.client.Dispatch("Outlook.Application")
    newMail = obj.CreateItem(olMailItem)
    newMail.Subject = subject
    newMail.Body = text
    # newMail.HTMLBody  = htmltext
    newMail.To = recipient
    # newMail.CC = "mneil@netapp.com"
    # newMail.BCC = "mneil@netapp.com"
    # attachment1 = "c:\\mypic.jpg"
    # newMail.Attachments.Add(attachment1)
    newMail.Send()    

# send_mail_via_com("email text", "email subject", "mneil@netapp.com")


def email_delivery_contacts(logistics_file, send_emails=False):

    # check for file
    if not os.path.isfile(logistics_file):
        print("Could not find quote file " + logistics_file, file=sys.stderr)
        return

    try:
        with open("signature.txt") as f:
            signature = f.read()
            print("Using following signature from signature.txt:", file=sys.stderr)
            print(signature, file=sys.stderr)

    except FileNotFoundError:
        with open("signature.txt", 'w') as f:
            print("NetApp", file=f)
            signature = "NetApp"
            print("Created signature.txt", file=sys.stderr)

    # read quote from quote.xlsx
    wb = openpyxl.load_workbook(logistics_file, read_only=True)
    sheet = wb.active
    rows = sheet.rows

    # column headings in header_row
    header_row = [cell.value for cell in next(rows)]

    while header_row[0] != "Serial Number Owner Name":
        header_row = [cell.value for cell in next(rows)]

    logistics_rows = []
    emails = set()
    sites = {}
    email_addresses = {}
    # put remaining rows in logistics_rows

    wb_out = openpyxl.Workbook(write_only=True)
    ws = wb_out.create_sheet()
    ws.append(header_row)

    for row in rows:
        record = {}

        # store each row in record
        for key, cell in zip(header_row, row):

            if cell.value is None:
                record[key] = ''
            elif cell.data_type == 's':
                # strip extra spaces from strings
                record[key] = cell.value.strip()
            else:
                # store everything else (numbers and dates) unchanged
                record[key] = cell.value

        # add row/record to logistics_rows
        logistics_rows.append(record)
        email_text = email_part_1 + \
            record['Installed At Site Name'] + \
            email_part_2 + \
            "Name: " + record['Delivery Contact Name'] + '\n' + \
            "Phone: " + record['Delivery Contact Phone'] + '\n' + \
            "Email: " + record['Delivery Contact eMail'] + '\n' + \
            record['Logistics Ship To Address Party Name 1'] + '\n' + \
            record['Logistics Address'] + '\n' + \
            record['Logisitcs City'] + '\n' + \
            record['Logistics State/Province'] + '\n' + \
            record['Logistic Postal Code'] + '\n' + \
            record['Logistics Country'] + '\n\n' + \
            "Receiving Hours:\n " + record['Goods Receiving Hour'] + '\n\n' + \
            "Service Report To Address:" + '\n\n' + \
            record['Service Report To Address'] + '\n' + \
            record['Service Report To City'] + '\n' +  \
            record['Service Report To Region'] + '\n' + \
            record['Service Report To Postal Code'] + '\n' + \
            record['Service Report To Country'] + \
            email_part_3 + signature

        if "UNKNOWN" not in record['Delivery Contact eMail'].upper() and email_text.lower() not in emails:
            ws.append([cell.value for cell in row])
            emails.add(email_text.lower())
            if debug_it:
                print(email_text, file=sys.stderr)
            if record['Installed At Site Name'] not in sites:
                sites[record['Installed At Site Name']] = 0
            sites[record['Installed At Site Name']] += 1
            if record["Delivery Contact eMail"].lower() not in email_addresses:
                email_addresses[record["Delivery Contact eMail"].lower()] = 0
            email_addresses[record["Delivery Contact eMail"].lower()] += 1

            if send_emails:
                msg_box_text = "Emails sent to " + record["Delivery Contact eMail"].lower() + ": " + str(email_addresses[record["Delivery Contact eMail"].lower()]) + \
                    '\n\n' + "Emails sent for site " + record['Installed At Site Name'] + ": " + str(sites[record['Installed At Site Name']]) + '\n\n' + \
                    email_text
                my_response = askyesnocancel("Click Yes to send email, No to skip, or Cancel to abort remaining", msg_box_text)
                if my_response:
                    send_mail_via_com(email_text, "Please respond:Data Center address verification request", "mneil@netapp.com")
                    sleep(2)
                elif my_response is None:
                    print("Emails aborted.", file=sys.stderr)
                    send_emails = False
        elif "UNKNOWN" in record['Delivery Contact eMail'].upper():
            ws.append([cell.value for cell in row])

    save_file_name = logistics_file.replace('.xlsx', '_scrubed.xlsx')

    try:
        wb_out.save(save_file_name)

        print("Done, opening " + save_file_name, file=sys.stderr)
        sleep(6)
        if platform.system() == "Darwin":
            # mac
            os.system("open " + save_file_name)
        else:
            # pc
            os.system("start " + save_file_name)
    except PermissionError:
        print("Can't write " + save_file_name + ", is the file open?", file=sys.stderr)

if __name__ == "__main__":
    print("Attempting to process logistics.xlsx", file=sys.stderr)
    email_delivery_contacts("logistics.xlsx")
